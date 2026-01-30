"""Представления для голосований."""

from __future__ import annotations

import csv
from io import BytesIO

from django.conf import settings
from django.contrib import messages
from django.contrib.admin.views.decorators import staff_member_required
from django.core.paginator import Paginator
from django.db import IntegrityError, models, transaction
from django.db.models import Count
from django.http import HttpRequest, HttpResponse, HttpResponseForbidden
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone

from .forms import ChoiceForm, PollCreateForm, PollVoteForm, QuestionForm
from .models import Choice, Poll, Question, Vote, VoteAnswer, Workplace
from .utils import get_client_ip

try:
    from openpyxl import Workbook
except ImportError:  # pragma: no cover
    Workbook = None

try:
    import xlwt
except ImportError:  # pragma: no cover
    xlwt = None


POLL_RESULT_COLORS = [
    "polls-color-emerald",
    "polls-color-blue",
    "polls-color-violet",
    "polls-color-amber",
    "polls-color-rose",
    "polls-color-teal",
]


def poll_list(request: HttpRequest) -> HttpResponse:
    if request.user.is_staff:
        polls = Poll.objects.all().order_by("-start_at", "-created_at")
    else:
        polls = Poll.objects.exclude(status=Poll.Status.DRAFT).order_by("-start_at", "-created_at")
    return render(request, "polls/poll_list.html", {"polls": polls, "now": timezone.now()})


@staff_member_required
def poll_create(request: HttpRequest) -> HttpResponse:
    if request.method == "POST":
        form = PollCreateForm(request.POST)
        if form.is_valid():
            poll = form.save()
            messages.success(request, "Голосование создано. Добавьте вопросы в конструкторе.")
            return redirect("polls:poll_builder", poll_id=poll.pk)
    else:
        form = PollCreateForm()

    return render(request, "polls/poll_create.html", {"form": form})


@staff_member_required
def poll_builder(request: HttpRequest, poll_id: int) -> HttpResponse:
    poll = get_object_or_404(Poll, pk=poll_id)
    if request.method == "POST":
        form = PollCreateForm(request.POST, instance=poll)
        if form.is_valid():
            form.save()
            messages.success(request, "Параметры голосования сохранены.")
            return redirect("polls:poll_builder", poll_id=poll.pk)
    else:
        form = PollCreateForm(instance=poll)

    questions = poll.questions.prefetch_related("choices").order_by("order", "id")

    return render(
        request,
        "polls/poll_builder.html",
        {
            "poll": poll,
            "form": form,
            "questions": questions,
        },
    )


@staff_member_required
def poll_question_create(request: HttpRequest, poll_id: int) -> HttpResponse:
    poll = get_object_or_404(Poll, pk=poll_id)
    question = Question(poll=poll)
    return _handle_question_form(request, poll, question)


@staff_member_required
def poll_question_edit(request: HttpRequest, poll_id: int, question_id: int) -> HttpResponse:
    poll = get_object_or_404(Poll, pk=poll_id)
    question = get_object_or_404(Question, pk=question_id, poll=poll)
    return _handle_question_form(request, poll, question)


@staff_member_required
def poll_question_delete(request: HttpRequest, poll_id: int, question_id: int) -> HttpResponse:
    poll = get_object_or_404(Poll, pk=poll_id)
    question = get_object_or_404(Question, pk=question_id, poll=poll)
    if request.method == "POST":
        question.delete()
        messages.success(request, "Вопрос удалён.")
    return redirect("polls:poll_builder", poll_id=poll.pk)


def _get_poll(slug_or_id: str) -> Poll:
    poll = Poll.objects.filter(slug=slug_or_id).first()
    if poll:
        return poll
    if slug_or_id.isdigit():
        return get_object_or_404(Poll, pk=int(slug_or_id))
    return get_object_or_404(Poll, slug=slug_or_id)


def poll_detail(request: HttpRequest, slug_or_id: str) -> HttpResponse:
    poll = _get_poll(slug_or_id)
    now = timezone.now()
    client_ip = get_client_ip(request)

    if not client_ip:
        return HttpResponseForbidden("Не удалось определить IP-адрес.")

    workplace = Workplace.objects.filter(ip_address=client_ip, is_active=True).first()
    eligible_workplaces = poll.audience_queryset()

    if not eligible_workplaces.filter(ip_address=client_ip).exists():
        return render(
            request,
            "polls/poll_unavailable.html",
            {
                "poll": poll,
                "status_message": "Ваше рабочее место не входит в аудиторию голосования.",
            },
            status=403,
        )

    if poll.status in {Poll.Status.DRAFT, Poll.Status.CLOSED, Poll.Status.ARCHIVED}:
        return render(
            request,
            "polls/poll_unavailable.html",
            {
                "poll": poll,
                "status_message": "Голосование недоступно.",
            },
            status=403,
        )

    if poll.start_at and now < poll.start_at:
        return render(
            request,
            "polls/poll_unavailable.html",
            {
                "poll": poll,
                "status_message": "Голосование ещё не началось.",
            },
        )

    if poll.end_at and now > poll.end_at:
        return render(
            request,
            "polls/poll_unavailable.html",
            {
                "poll": poll,
                "status_message": "Голосование завершено.",
                "show_results": _can_view_results(request, poll),
            },
        )

    existing_vote = Vote.objects.filter(poll=poll, voter_ip=client_ip).first()

    if existing_vote and poll.vote_policy == Poll.VotePolicy.SINGLE:
        return render(
            request,
            "polls/poll_unavailable.html",
            {
                "poll": poll,
                "status_message": "Вы уже голосовали. Повторное голосование недоступно.",
                "show_results": _can_view_results(request, poll),
            },
            status=403,
        )

    questions = poll.questions.all().order_by("order", "id").prefetch_related("choices")

    if request.method == "POST":
        form = PollVoteForm(request.POST, poll=poll)
        if form.is_valid():
            try:
                with transaction.atomic():
                    vote = _save_vote(
                        poll=poll,
                        form=form,
                        client_ip=client_ip,
                        workplace=workplace,
                        existing_vote=existing_vote,
                    )
                    _save_answers(vote, form)
            except IntegrityError:
                messages.error(request, "Не удалось сохранить голос. Попробуйте ещё раз.")
            else:
                messages.success(request, "Ваш голос учтён.")
                return redirect("polls:poll_thanks", poll_id=poll.pk)
    else:
        form = PollVoteForm(poll=poll)

    question_blocks = []
    for question in questions:
        field_name = f"question_{question.pk}"
        try:
            field = form[field_name]
        except KeyError:
            field = None
        question_blocks.append({"question": question, "field": field})

    return render(
        request,
        "polls/poll_detail.html",
        {
            "poll": poll,
            "form": form,
            "workplace": workplace,
            "questions": questions,
            "question_blocks": question_blocks,
            "show_missing_fields": settings.DEBUG or request.user.is_staff,
        },
    )


def _save_vote(
    *,
    poll: Poll,
    form: PollVoteForm,
    client_ip: str,
    workplace: Workplace | None,
    existing_vote: Vote | None,
) -> Vote:
    full_name = form.cleaned_data.get("full_name", "")

    if existing_vote and poll.vote_policy == Poll.VotePolicy.REVOTE_UNTIL_END and not poll.has_ended:
        existing_vote.full_name = full_name
        existing_vote.workplace = workplace
        existing_vote.save(update_fields=["full_name", "workplace", "updated_at"])
        existing_vote.answers.all().delete()
        return existing_vote

    return Vote.objects.create(
        poll=poll,
        voter_ip=client_ip,
        workplace=workplace,
        full_name=full_name,
    )


def _save_answers(vote: Vote, form: PollVoteForm) -> None:
    for question in vote.poll.questions.all():
        field_name = f"question_{question.pk}"
        value = form.cleaned_data.get(field_name)
        if value in (None, ""):
            continue
        if question.type == Question.QuestionType.TEXT:
            VoteAnswer.objects.create(vote=vote, question=question, answer_text=value)
        elif question.type == Question.QuestionType.MULTI_CHOICE:
            for choice in value:
                VoteAnswer.objects.create(vote=vote, question=question, choice=choice)
        else:
            VoteAnswer.objects.create(vote=vote, question=question, choice=value)


def poll_thanks(request: HttpRequest, poll_id: int) -> HttpResponse:
    poll = get_object_or_404(Poll, pk=poll_id)
    return render(request, "polls/poll_thanks.html", {"poll": poll})


def _can_view_results(request: HttpRequest, poll: Poll) -> bool:
    return request.user.is_staff or poll.show_results_to_users or poll.has_ended


def poll_results(request: HttpRequest, poll_id: int) -> HttpResponse:
    poll = get_object_or_404(Poll, pk=poll_id)
    if not _can_view_results(request, poll):
        return HttpResponseForbidden("Результаты недоступны.")

    result_blocks = []
    text_pages = {}
    text_page_links = {}

    for question in poll.questions.prefetch_related("choices"):
        answers_qs = VoteAnswer.objects.filter(question=question)
        total_votes = answers_qs.values("vote_id").distinct().count()

        if question.type == Question.QuestionType.TEXT:
            paginator = Paginator(answers_qs.exclude(answer_text=""), 20)
            page_number = request.GET.get(f"q{question.pk}_page", 1)
            answers_page = paginator.get_page(page_number)
            text_pages[question.pk] = answers_page
            page_links = {"prev": None, "next": None}
            if answers_page.has_previous():
                prev_query = request.GET.copy()
                prev_query[f"q{question.pk}_page"] = answers_page.previous_page_number()
                page_links["prev"] = f"?{prev_query.urlencode()}"
            if answers_page.has_next():
                next_query = request.GET.copy()
                next_query[f"q{question.pk}_page"] = answers_page.next_page_number()
                page_links["next"] = f"?{next_query.urlencode()}"
            text_page_links[question.pk] = page_links
            result_blocks.append(
                {
                    "question": question,
                    "total": total_votes,
                    "type": "text",
                }
            )
            continue

        choice_counts = (
            answers_qs.filter(choice__isnull=False)
            .values("choice")
            .annotate(total=Count("choice"))
        )
        count_map = {item["choice"]: item["total"] for item in choice_counts}

        choices_data = []
        for idx, choice in enumerate(question.choices.all()):
            count = count_map.get(choice.pk, 0)
            percent = (count / total_votes * 100) if total_votes else 0
            color_class = POLL_RESULT_COLORS[idx % len(POLL_RESULT_COLORS)]
            choices_data.append(
                {
                    "choice": choice,
                    "count": count,
                    "percent": percent,
                    "color_class": color_class,
                }
            )

        result_blocks.append(
            {
                "question": question,
                "total": total_votes,
                "choices": choices_data,
                "type": "choices",
            }
        )

    turnout = _build_turnout(poll)

    return render(
        request,
        "polls/poll_results.html",
        {
            "poll": poll,
            "result_blocks": result_blocks,
            "text_pages": text_pages,
            "text_page_links": text_page_links,
            "turnout": turnout,
        },
    )


def poll_results_csv(request: HttpRequest, poll_id: int) -> HttpResponse:
    poll = get_object_or_404(Poll, pk=poll_id)
    if not _can_view_results(request, poll):
        return HttpResponseForbidden("Результаты недоступны.")

    response = HttpResponse(content_type="text/csv; charset=utf-8")
    response["Content-Disposition"] = f'attachment; filename="poll_{poll_id}_results.csv"'
    response.write("\ufeff")

    writer = csv.writer(response)
    _write_results_rows(writer, poll)

    return response


def poll_results_xlsx(request: HttpRequest, poll_id: int) -> HttpResponse:
    poll = get_object_or_404(Poll, pk=poll_id)
    if not _can_view_results(request, poll):
        return HttpResponseForbidden("Результаты недоступны.")
    if Workbook is None:  # pragma: no cover
        return HttpResponse("XLSX недоступен на сервере.", status=501)

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Результаты"
    _write_results_rows(worksheet.append, poll)

    content = BytesIO()
    workbook.save(content)
    content.seek(0)

    response = HttpResponse(
        content.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="poll_{poll_id}_results.xlsx"'
    return response


def poll_results_xls(request: HttpRequest, poll_id: int) -> HttpResponse:
    poll = get_object_or_404(Poll, pk=poll_id)
    if not _can_view_results(request, poll):
        return HttpResponseForbidden("Результаты недоступны.")
    if xlwt is None:  # pragma: no cover
        return HttpResponse("XLS недоступен на сервере.", status=501)

    workbook = xlwt.Workbook()
    worksheet = workbook.add_sheet("Результаты")
    _write_results_rows(lambda row: _xls_append_row(worksheet, row), poll)

    content = BytesIO()
    workbook.save(content)
    content.seek(0)

    response = HttpResponse(content.getvalue(), content_type="application/vnd.ms-excel")
    response["Content-Disposition"] = f'attachment; filename="poll_{poll_id}_results.xls"'
    return response


@staff_member_required
def poll_results_people(request: HttpRequest, poll_id: int) -> HttpResponse:
    poll = get_object_or_404(Poll, pk=poll_id)
    votes = Vote.objects.filter(poll=poll).select_related("workplace").order_by("-created_at")
    query = request.GET.get("q", "").strip()
    if query:
        votes = votes.filter(
            models.Q(full_name__icontains=query)
            | models.Q(voter_ip__icontains=query)
            | models.Q(workplace__label__icontains=query)
        )

    return render(
        request,
        "polls/poll_results_people.html",
        {
            "poll": poll,
            "votes": votes,
            "query": query,
        },
    )


@staff_member_required
def poll_results_people_csv(request: HttpRequest, poll_id: int) -> HttpResponse:
    poll = get_object_or_404(Poll, pk=poll_id)
    response = HttpResponse(content_type="text/csv; charset=utf-8")
    response["Content-Disposition"] = f'attachment; filename="poll_{poll_id}_people.csv"'
    response.write("\ufeff")
    writer = csv.writer(response)
    _write_people_rows(writer, poll)
    return response


@staff_member_required
def poll_results_people_xlsx(request: HttpRequest, poll_id: int) -> HttpResponse:
    poll = get_object_or_404(Poll, pk=poll_id)
    if Workbook is None:  # pragma: no cover
        return HttpResponse("XLSX недоступен на сервере.", status=501)

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Ответы"
    _write_people_rows(worksheet.append, poll)

    content = BytesIO()
    workbook.save(content)
    content.seek(0)

    response = HttpResponse(
        content.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="poll_{poll_id}_people.xlsx"'
    return response


@staff_member_required
def poll_results_people_xls(request: HttpRequest, poll_id: int) -> HttpResponse:
    poll = get_object_or_404(Poll, pk=poll_id)
    if xlwt is None:  # pragma: no cover
        return HttpResponse("XLS недоступен на сервере.", status=501)

    workbook = xlwt.Workbook()
    worksheet = workbook.add_sheet("Ответы")
    _write_people_rows(lambda row: _xls_append_row(worksheet, row), poll)

    content = BytesIO()
    workbook.save(content)
    content.seek(0)

    response = HttpResponse(content.getvalue(), content_type="application/vnd.ms-excel")
    response["Content-Disposition"] = f'attachment; filename="poll_{poll_id}_people.xls"'
    return response


@staff_member_required
def poll_results_person(request: HttpRequest, poll_id: int, vote_id: int) -> HttpResponse:
    poll = get_object_or_404(Poll, pk=poll_id)
    vote = get_object_or_404(Vote, pk=vote_id, poll=poll)
    questions = list(poll.questions.order_by("order", "id").prefetch_related("choices"))
    answers = VoteAnswer.objects.filter(vote=vote).select_related("choice", "question")
    answers_map: dict[int, list[VoteAnswer]] = {}
    for answer in answers:
        answers_map.setdefault(answer.question_id, []).append(answer)

    return render(
        request,
        "polls/poll_results_person.html",
        {
            "poll": poll,
            "vote": vote,
            "questions": questions,
            "answers_map": answers_map,
        },
    )


@staff_member_required
def poll_turnout(request: HttpRequest, poll_id: int) -> HttpResponse:
    poll = get_object_or_404(Poll, pk=poll_id)
    turnout = _build_turnout(poll)

    return render(
        request,
        "polls/poll_turnout.html",
        {
            "poll": poll,
            "turnout": turnout,
        },
    )


@staff_member_required
def poll_turnout_csv(request: HttpRequest, poll_id: int) -> HttpResponse:
    poll = get_object_or_404(Poll, pk=poll_id)
    turnout = _build_turnout(poll)

    response = HttpResponse(content_type="text/csv; charset=utf-8")
    response["Content-Disposition"] = f'attachment; filename="poll_{poll_id}_turnout.csv"'
    response.write("\ufeff")

    writer = csv.writer(response)
    writer.writerow(["Голосование", poll.title])
    writer.writerow(["Всего аудитории", turnout["eligible_count"]])
    writer.writerow(["Проголосовали", turnout["voted_count"]])
    writer.writerow(["Не проголосовали", turnout["not_voted_count"]])

    writer.writerow([])
    writer.writerow(["Проголосовали (IP)"])
    for workplace in turnout["voted_workplaces"]:
        writer.writerow([workplace.ip_address, workplace.label, workplace.department])

    writer.writerow([])
    writer.writerow(["Не проголосовали (IP)"])
    for workplace in turnout["not_voted_workplaces"]:
        writer.writerow([workplace.ip_address, workplace.label, workplace.department])

    return response


def _build_turnout(poll: Poll) -> dict:
    eligible_workplaces = poll.audience_queryset()
    voted_ips = Vote.objects.filter(poll=poll).values_list("voter_ip", flat=True)
    voted_workplaces = eligible_workplaces.filter(ip_address__in=voted_ips)
    not_voted_workplaces = eligible_workplaces.exclude(ip_address__in=voted_ips)

    return {
        "eligible_count": eligible_workplaces.count(),
        "voted_count": voted_workplaces.count(),
        "not_voted_count": not_voted_workplaces.count(),
        "voted_workplaces": voted_workplaces,
        "not_voted_workplaces": not_voted_workplaces,
    }


def _handle_question_form(request: HttpRequest, poll: Poll, question: Question) -> HttpResponse:
    from django.forms import inlineformset_factory

    ChoiceFormSet = inlineformset_factory(
        Question,
        Choice,
        form=ChoiceForm,
        extra=2,
        can_delete=True,
    )

    if request.method == "POST":
        form = QuestionForm(request.POST, request.FILES, instance=question)
        saved_question = question
        if form.is_valid():
            saved_question = form.save(commit=False)
            saved_question.poll = poll
            saved_question.save()

            if saved_question.type == Question.QuestionType.TEXT:
                Choice.objects.filter(question=saved_question).delete()
                messages.success(request, "Вопрос сохранён.")
                return redirect("polls:poll_builder", poll_id=poll.pk)

            formset = ChoiceFormSet(request.POST, instance=saved_question)
            if formset.is_valid():
                formset.save()
                has_choices = saved_question.choices.exists()
                if not has_choices:
                    formset._non_form_errors = formset.error_class(
                        ["Добавьте хотя бы один вариант ответа."]
                    )
                else:
                    messages.success(request, "Вопрос сохранён.")
                    return redirect("polls:poll_builder", poll_id=poll.pk)
        formset = ChoiceFormSet(request.POST, instance=saved_question)
    else:
        form = QuestionForm(instance=question)
        formset = ChoiceFormSet(instance=question)

    return render(
        request,
        "polls/poll_question_form.html",
        {
            "poll": poll,
            "question": question if question.pk else None,
            "form": form,
            "formset": formset,
        },
    )


def _write_results_rows(write_row, poll: Poll) -> None:
    write_row(["Голосование", poll.title])
    for question in poll.questions.prefetch_related("choices"):
        write_row([])
        write_row(["Вопрос", question.text])
        answers_qs = VoteAnswer.objects.filter(question=question)
        total_votes = answers_qs.values("vote_id").distinct().count()

        if question.type == Question.QuestionType.TEXT:
            write_row(["Тип", "Текстовые ответы"])
            for answer in answers_qs.exclude(answer_text=""):
                write_row([answer.answer_text])
            continue

        write_row(["Всего ответов", total_votes])
        choice_counts = (
            answers_qs.filter(choice__isnull=False)
            .values("choice")
            .annotate(total=Count("choice"))
        )
        count_map = {item["choice"]: item["total"] for item in choice_counts}
        for choice in question.choices.all():
            count = count_map.get(choice.pk, 0)
            percent = (count / total_votes * 100) if total_votes else 0
            write_row([choice.text, count, f"{percent:.2f}%"])


def _write_people_rows(write_row, poll: Poll) -> None:
    questions = list(poll.questions.order_by("order", "id"))
    header = ["IP", "Рабочее место", "Подразделение", "ФИО", "Дата"]
    header.extend([question.text for question in questions])
    write_row(header)

    answers = (
        VoteAnswer.objects.filter(vote__poll=poll)
        .select_related("vote", "choice", "question", "vote__workplace")
        .order_by("vote__created_at")
    )
    answers_map: dict[int, dict[int, list[VoteAnswer]]] = {}
    for answer in answers:
        answers_map.setdefault(answer.vote_id, {}).setdefault(answer.question_id, []).append(answer)

    for vote in Vote.objects.filter(poll=poll).select_related("workplace").order_by("created_at"):
        row = [
            vote.voter_ip,
            vote.workplace.label if vote.workplace else "",
            vote.workplace.department if vote.workplace else "",
            vote.full_name,
            timezone.localtime(vote.created_at).strftime("%d.%m.%Y %H:%M"),
        ]
        for question in questions:
            answer_list = answers_map.get(vote.id, {}).get(question.id, [])
            if question.type == Question.QuestionType.TEXT:
                cell = ", ".join(a.answer_text for a in answer_list if a.answer_text)
            else:
                cell = ", ".join(a.choice.text for a in answer_list if a.choice)
            row.append(cell)
        write_row(row)


def _xls_append_row(worksheet, row: list) -> None:
    row_index = getattr(worksheet, "_osiris_row", 0)
    for col_index, value in enumerate(row):
        worksheet.write(row_index, col_index, value)
    setattr(worksheet, "_osiris_row", row_index + 1)
